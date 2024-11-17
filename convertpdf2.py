import pdfplumber
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

def pdf_to_excel_with_borders(pdf_path, excel_path):
    """Extract tables from PDF and save them into one Excel sheet with custom borders for first and second row, and center text."""
    with pdfplumber.open(pdf_path) as pdf:
        all_tables = []

        for page in pdf.pages:
            # Extract tables from the page
            tables = page.extract_tables()

            # Only process non-empty tables
            for table in tables:
                if table and any(cell.strip() for row in table for cell in row):
                    # Convert each non-empty table to DataFrame and append to list
                    df = pd.DataFrame(table[1:], columns=table[0])
                    all_tables.append(df)

        # Combine all tables into one DataFrame
        combined_df = pd.concat(all_tables, ignore_index=True)

        # Save the combined DataFrame to Excel
        combined_df.to_excel(excel_path, index=False)

    # Load the saved Excel file with openpyxl to add borders, adjust column widths, and center text
    wb = load_workbook(excel_path)
    ws = wb.active

    # Define the border style for regular cells (thin)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define the border style for first and second row (thicker border)
    thicker_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Define the alignment style (centered)
    centered_alignment = Alignment(horizontal='center', vertical='center')

    # Apply the thicker border and center text to the first and second rows
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thicker_border
            cell.alignment = centered_alignment

    # Apply regular borders to all other cells and center text
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_style

    # Auto-fit column widths based on the longest item in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (A, B, C, ...)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the modified Excel file with borders, auto-fit columns, and centered text
    wb.save(excel_path)
    print(f"Converted {pdf_path} to {excel_path} with custom borders and centered text for the first and second rows.")

def convert_all_pdfs_in_folder(folder_path):
    """Converts all PDF files in the specified folder to one Excel file each with custom borders and centered text for first and second rows."""
    for filename in os.listdir(folder_path):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(folder_path, filename)
            excel_path = os.path.join(folder_path, f"{os.path.splitext(filename)[0]}.xlsx")
            pdf_to_excel_with_borders(pdf_path, excel_path)
          
# Replace 'path/to/your/folder' with the path to the folder containing your PDFs
folder_path = '\\input'
convert_all_pdfs_in_folder(folder_path)
